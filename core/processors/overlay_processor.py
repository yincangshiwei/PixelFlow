"""
图片叠加处理器 —— 在图片上叠加文本和图片元素
支持：固定文本、Excel列数据、图片名、外部图片叠加
"""
import os
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QLabel, QSpinBox, QComboBox, QPushButton, QColorDialog,
    QListWidget, QListWidgetItem, QLineEdit, QCheckBox,
    QFileDialog, QDoubleSpinBox, QTextEdit, QMessageBox,
    QStackedWidget
)
from PySide6.QtCore import Qt, QRectF
from PySide6.QtGui import QColor, QPainter, QPen, QBrush, QFont

from core.base_processor import BaseProcessor, register_processor
from core.image_processor import hex_to_rgba
import config


class OverlayElement:
    """叠加元素基类"""
    def __init__(self, element_type, x, y, name=''):
        self.element_type = element_type  # 'text' or 'image'
        self.x = x  # X坐标（像素值）
        self.y = y  # Y坐标（像素值）
        self.name = name  # 用户自定义名称


class TextElement(OverlayElement):
    """文本元素"""
    def __init__(self, x=50, y=50, source='fixed', content='', 
                 font_size=24, font_family='Microsoft YaHei', bold=False, 
                 color='#FFFFFF', excel_file='', match_column=0, data_column=0, 
                 excel_row_start=2, name=''):
        super().__init__('text', x, y, name)
        self.source = source  # 'fixed', 'excel', 'filename'
        self.content = content  # 固定文本内容
        self.font_size = font_size
        self.font_family = font_family
        self.bold = bold
        self.color = color
        self.excel_file = excel_file
        self.match_column = match_column  # 匹配列（图片名称所在列，从1开始）
        self.data_column = data_column    # 数据列（要读取的文本数据列，从1开始）
        self.excel_row_start = excel_row_start  # 数据起始行


class ImageElement(OverlayElement):
    """图片元素"""
    def __init__(self, x=100, y=100, image_path='', width=200, height=200, name=''):
        super().__init__('image', x, y, name)
        self.image_path = image_path
        self.width = width  # 宽度（像素值）
        self.height = height  # 高度（像素值）


class ElementListItem(QListWidgetItem):
    """元素列表项"""
    def __init__(self, element, index):
        super().__init__()
        self.element = element
        self.index = index
        self.update_text()
    
    def update_text(self):
        label = self.element.name if self.element.name else ""
        if self.element.element_type == 'text':
            self.setText(f"📝 文本 {self.index + 1} | {label}")
        else:
            self.setText(f"🖼️ 图片 {self.index + 1} | {label}")


class GridPositionWidget(QWidget):
    """3x3 宫格坐标定位器 — 点击格点自动填充 X/Y 坐标值"""

    # 9宫格定义：(箭头符号, X比例, Y比例)
    CELLS = [
        ("↖", 0.0, 0.0),  # 左上
        ("↑", 0.5, 0.0),  # 中上
        ("↗", 1.0, 0.0),  # 右上
        ("←", 0.0, 0.5),  # 左
        ("⊙", 0.5, 0.5),  # 中
        ("→", 1.0, 0.5),  # 右
        ("↙", 0.0, 1.0),  # 左下
        ("↓", 0.5, 1.0),  # 中下
        ("↘", 1.0, 1.0),  # 右下
    ]

    def __init__(self, x_spin, y_spin, processor, element_type='text',
                 overlay_w_spin=None, overlay_h_spin=None, parent=None):
        super().__init__(parent)
        self._x_spin = x_spin
        self._y_spin = y_spin
        self._processor = processor  # 用于获取基础图片尺寸
        self._element_type = element_type  # 'text' 或 'image'
        self._overlay_w_spin = overlay_w_spin
        self._overlay_h_spin = overlay_h_spin
        self._hovered_cell = -1

        self.setFixedSize(90, 90)
        self.setMouseTracking(True)
        self.setCursor(Qt.PointingHandCursor)
        self.setToolTip("点击格点自动计算并填充 X/Y 坐标\n"
                        "基于左侧列表选中图片的尺寸定位")

    def _get_base_size(self):
        """获取基础图片尺寸（优先从处理器读取选中图片的实际尺寸）"""
        bw = getattr(self._processor, '_base_img_width', 1920)
        bh = getattr(self._processor, '_base_img_height', 1080)
        return bw, bh

    def _cell_at(self, pos):
        """根据鼠标坐标计算所在单元格索引（0~8），不在格内返回 -1"""
        cw = self.width() / 3.0
        ch = self.height() / 3.0
        col = int(pos.x() / cw)
        row = int(pos.y() / ch)
        if 0 <= col < 3 and 0 <= row < 3:
            return row * 3 + col
        return -1

    def _apply_position(self, idx):
        """根据格点索引计算并设置 X/Y 坐标值"""
        if not getattr(self._processor, '_has_base_image', False):
            QMessageBox.information(
                self._processor._panel, "提示",
                "请先在左侧列表选中一张图片，\n宫格定位需要知道底图尺寸。"
            )
            return
        _, px, py = self.CELLS[idx]
        bw, bh = self._get_base_size()
        x = int(bw * px)
        y = int(bh * py)

        # 图片元素：根据叠加图片自身尺寸对居中/靠右/靠下位置进行偏移
        if self._element_type == 'image':
            ow = self._overlay_w_spin.value() if self._overlay_w_spin else 0
            oh = self._overlay_h_spin.value() if self._overlay_h_spin else 0
            if px == 0.5:
                x -= ow // 2
            elif px == 1.0:
                x -= ow
            if py == 0.5:
                y -= oh // 2
            elif py == 1.0:
                y -= oh

        self._x_spin.setValue(max(0, x))
        self._y_spin.setValue(max(0, y))

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        w = self.width()
        h = self.height()
        cw = w / 3.0
        ch = h / 3.0

        # 整体背景
        painter.setPen(Qt.NoPen)
        painter.setBrush(QBrush(QColor(18, 18, 40, 180)))
        painter.drawRoundedRect(0, 0, w, h, 8, 8)

        # 绘制 3x3 格线
        pen = QPen(QColor(100, 110, 170, 60), 1)
        painter.setPen(pen)
        painter.setBrush(Qt.NoBrush)
        for i in range(4):
            painter.drawLine(int(i * cw), 0, int(i * cw), h)
            painter.drawLine(0, int(i * ch), w, int(i * ch))

        # 绘制箭头符号
        arrow_font = QFont("Segoe UI Symbol", 12)
        painter.setFont(arrow_font)

        for idx, (arrow, _, _) in enumerate(self.CELLS):
            row = idx // 3
            col = idx % 3
            rect = QRectF(col * cw, row * ch, cw, ch)

            # 悬停高亮
            if idx == self._hovered_cell:
                painter.setPen(Qt.NoPen)
                painter.setBrush(QBrush(QColor(91, 138, 245, 80)))
                painter.drawRoundedRect(rect.adjusted(2, 2, -2, -2), 5, 5)
                painter.setPen(QColor(200, 210, 255))
            else:
                painter.setPen(QColor(150, 160, 190))

            painter.drawText(rect, Qt.AlignCenter, arrow)

        painter.end()

    def mouseMoveEvent(self, event):
        cell = self._cell_at(event.pos())
        if cell != self._hovered_cell:
            self._hovered_cell = cell
            self.update()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            cell = self._cell_at(event.pos())
            if cell >= 0:
                self._apply_position(cell)

    def leaveEvent(self, event):
        if self._hovered_cell != -1:
            self._hovered_cell = -1
            self.update()


@register_processor
class OverlayProcessor(BaseProcessor):
    """图片叠加处理器"""

    name = "图片叠加"
    description = "在图片上叠加文本和图片元素，支持Excel数据、图片名称等数据源"
    icon = "🎨"
    preset_id = "image_overlay"

    def __init__(self):
        self._panel: QWidget | None = None
        self._elements: list[OverlayElement] = []
        self._list_widget: QListWidget | None = None
        self._config_stack: QStackedWidget | None = None  # 使用 QStackedWidget 管理配置面板
        self._current_file_path = ""  # 当前处理的图片路径
        self._font_cache: dict[str, ImageFont.FreeTypeFont] = {}  # 字体缓存，避免重复加载导致内存泄漏
        self._base_img_width = 1920   # 选中底图的宽度（像素），用于宫格定位
        self._base_img_height = 1080  # 选中底图的高度（像素），用于宫格定位
        self._has_base_image = False  # 是否已选中底图

    def set_base_image_size(self, w: int, h: int):
        """由主窗口调用，同步当前选中图片的尺寸（像素）"""
        if w > 0 and h > 0:
            self._base_img_width = w
            self._base_img_height = h
            self._has_base_image = True

    def _clear_font_cache(self):
        """清理字体缓存，释放内存"""
        for font in self._font_cache.values():
            try:
                del font
            except:
                pass
        self._font_cache.clear()

    def create_panel(self, parent=None) -> QWidget:
        self._panel = QWidget(parent)
        root = QVBoxLayout(self._panel)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(8)

        # ── 元素列表 + 配置（左右分栏）──
        grp_elements = QGroupBox("叠加元素")
        elem_layout = QHBoxLayout(grp_elements)
        elem_layout.setSpacing(10)

        # ── 左侧：元素列表 + emoji按钮 ──
        left_panel = QWidget()
        left_panel.setMaximumWidth(280)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(6)

        self._list_widget = QListWidget()
        self._list_widget.setMinimumHeight(140)
        self._list_widget.currentRowChanged.connect(self._on_element_selected)
        left_layout.addWidget(self._list_widget)

        # 操作按钮（emoji图标）
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(4)

        def _make_icon_btn(emoji, tooltip, slot):
            btn = QPushButton(emoji)
            btn.setToolTip(tooltip)
            btn.setStyleSheet(config.ICON_BTN_STYLE)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(slot)
            return btn

        btn_add_text = _make_icon_btn("📝", "添加文本", self._add_text_element)
        btn_layout.addWidget(btn_add_text)

        btn_add_image = _make_icon_btn("🖼", "添加图片", self._add_image_element)
        btn_layout.addWidget(btn_add_image)

        btn_delete = _make_icon_btn("🗑", "删除选中", self._delete_element)
        btn_layout.addWidget(btn_delete)

        btn_layout.addSpacing(4)

        btn_rename = _make_icon_btn("✏", "重命名", self._rename_element)
        btn_layout.addWidget(btn_rename)

        btn_layout.addSpacing(6)

        btn_move_up = _make_icon_btn("⬆", "上移", self._move_element_up)
        btn_layout.addWidget(btn_move_up)

        btn_move_down = _make_icon_btn("⬇", "下移", self._move_element_down)
        btn_layout.addWidget(btn_move_down)

        btn_layout.addStretch()
        left_layout.addLayout(btn_layout)

        elem_layout.addWidget(left_panel, stretch=2)

        # ── 右侧：元素配置面板 ──
        grp_config = QGroupBox("元素配置")
        config_root = QVBoxLayout(grp_config)
        config_root.setContentsMargins(8, 8, 8, 8)

        # 使用 QStackedWidget 管理不同的配置面板
        self._config_stack = QStackedWidget()

        # 默认显示提示页面
        hint_page = QWidget()
        hint_layout = QVBoxLayout(hint_page)
        hint_label = QLabel("请先添加元素，然后在左侧列表中选择一个元素进行配置")
        hint_label.setStyleSheet("color:#666e88;font-size:12px;font-style:italic;")
        hint_label.setAlignment(Qt.AlignCenter)
        hint_layout.addWidget(hint_label)
        self._config_stack.addWidget(hint_page)

        config_root.addWidget(self._config_stack)
        elem_layout.addWidget(grp_config, stretch=5)

        root.addWidget(grp_elements)

        # ── 输出格式 ──
        fmt_row = QHBoxLayout()
        fmt_row.addWidget(QLabel("输出格式:"))
        self.combo_fmt = QComboBox()
        self.combo_fmt.addItems(["png", "webp", "jpg"])
        self.combo_fmt.setStyleSheet(config.COMBOBOX_STYLE)
        fmt_row.addWidget(self.combo_fmt)
        fmt_row.addStretch()
        root.addLayout(fmt_row)

        root.addStretch()
        return self._panel

    def _add_text_element(self):
        """添加文本元素"""
        # 根据已有元素数量计算默认位置，避免重叠
        text_count = sum(1 for e in self._elements if e.element_type == 'text')
        default_y = 50 + (text_count * 50)  # 依次向下排列，避免重叠
        element = TextElement(x=50, y=default_y)
        self._elements.append(element)
        item = ElementListItem(element, len(self._elements) - 1)
        self._list_widget.addItem(item)
        self._list_widget.setCurrentRow(self._list_widget.count() - 1)

    def _add_image_element(self):
        """添加图片元素"""
        # 根据已有元素数量计算默认位置，避免重叠
        img_count = sum(1 for e in self._elements if e.element_type == 'image')
        default_y = 100 + (img_count * 250)  # 依次向下排列
        element = ImageElement(x=100, y=default_y)
        self._elements.append(element)
        item = ElementListItem(element, len(self._elements) - 1)
        self._list_widget.addItem(item)
        self._list_widget.setCurrentRow(self._list_widget.count() - 1)

    def _delete_element(self):
        """删除选中的元素"""
        row = self._list_widget.currentRow()
        if row < 0:
            return
        self._list_widget.takeItem(row)
        self._elements.pop(row)
        # 更新索引
        for i in range(self._list_widget.count()):
            item = self._list_widget.item(i)
            item.index = i
            item.update_text()

    def _rename_element(self):
        """重命名选中元素的名称（| 后面的文字）"""
        row = self._list_widget.currentRow()
        if row < 0:
            return
        from PySide6.QtWidgets import QInputDialog
        current_name = self._elements[row].name
        new_name, ok = QInputDialog.getText(
            self._panel, "重命名元素", "请输入新名称:",
            text=current_name
        )
        if ok:
            self._elements[row].name = new_name.strip()
            self._list_widget.item(row).update_text()

    def _move_element_up(self):
        """上移选中的元素"""
        row = self._list_widget.currentRow()
        if row <= 0:
            return
        # 交换列表项
        item = self._list_widget.takeItem(row)
        self._list_widget.insertItem(row - 1, item)
        self._list_widget.setCurrentRow(row - 1)
        # 交换元素
        self._elements[row], self._elements[row - 1] = self._elements[row - 1], self._elements[row]
        # 更新索引
        for i in range(self._list_widget.count()):
            self._list_widget.item(i).index = i
            self._list_widget.item(i).update_text()

    def _move_element_down(self):
        """下移选中的元素"""
        row = self._list_widget.currentRow()
        if row < 0 or row >= self._list_widget.count() - 1:
            return
        # 交换列表项
        item = self._list_widget.takeItem(row)
        self._list_widget.insertItem(row + 1, item)
        self._list_widget.setCurrentRow(row + 1)
        # 交换元素
        self._elements[row], self._elements[row + 1] = self._elements[row + 1], self._elements[row]
        # 更新索引
        for i in range(self._list_widget.count()):
            self._list_widget.item(i).index = i
            self._list_widget.item(i).update_text()

    def _on_element_selected(self, row):
        """选中元素时显示配置面板"""
        if row < 0:
            return
        
        # 先收集当前编辑元素的参数
        self._collect_current_element_options()
        
        element = self._elements[row]
        
        # 彻底清除旧配置页面（保留索引0的提示页面）
        self._clear_config_pages()
        
        # 创建新的配置页面
        if element.element_type == 'text':
            config_page = self._create_text_config_page(element)
        else:
            config_page = self._create_image_config_page(element)
        
        self._config_stack.addWidget(config_page)
        self._config_stack.setCurrentIndex(1)

    def _clear_config_pages(self):
        """彻底清除所有配置页面（保留索引0的提示页面）"""
        # 清除实例变量引用
        attrs_to_clear = [
            '_current_element', '_combo_source', '_widget_fixed_text', 
            '_widget_excel', '_widget_filename_hint', '_text_content',
            '_excel_file_input', '_excel_match_column', '_excel_data_column',
            '_excel_row_start',
            '_font_family', '_font_size', '_chk_bold', 
            '_text_color_btn', '_text_color_lbl', '_text_x', '_text_y',
            '_image_file_input', '_image_x', '_image_y', 
            '_image_width', '_image_height'
        ]
        for attr in attrs_to_clear:
            if hasattr(self, attr):
                delattr(self, attr)
        
        # 移除并删除所有配置页面（保留索引0）
        while self._config_stack.count() > 1:
            widget = self._config_stack.widget(1)
            self._config_stack.removeWidget(widget)
            widget.setParent(None)  # 断开父子关系
            widget.deleteLater()

    def _create_text_config_page(self, element: TextElement) -> QWidget:
        """创建文本元素配置页面"""
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(8)
        # 数据源选择
        src_layout = QHBoxLayout()
        src_layout.addWidget(QLabel("数据来源:"))
        combo_source = QComboBox()
        combo_source.addItem("固定文本", "fixed")
        combo_source.addItem("Excel列数据", "excel")
        combo_source.addItem("图片文件名称", "filename")
        combo_source.setCurrentIndex(combo_source.findData(element.source))
        combo_source.setStyleSheet(config.COMBOBOX_STYLE)
        src_layout.addWidget(combo_source)
        src_layout.addStretch()
        page_layout.addLayout(src_layout)

        # 固定文本内容
        self._widget_fixed_text = QWidget()
        fixed_layout = QVBoxLayout(self._widget_fixed_text)
        fixed_layout.setContentsMargins(0, 0, 0, 0)
        fixed_layout.addWidget(QLabel("文本内容:"))
        self._text_content = QTextEdit()
        self._text_content.setPlainText(element.content)
        self._text_content.setMaximumHeight(80)
        fixed_layout.addWidget(self._text_content)
        page_layout.addWidget(self._widget_fixed_text)

        # Excel配置
        self._widget_excel = QWidget()
        excel_layout = QVBoxLayout(self._widget_excel)
        excel_layout.setContentsMargins(0, 0, 0, 0)
        
        # Excel文件选择
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Excel文件:"))
        self._excel_file_input = QLineEdit()
        self._excel_file_input.setText(element.excel_file)
        self._excel_file_input.setPlaceholderText("选择Excel文件...")
        file_layout.addWidget(self._excel_file_input)
        btn_browse_excel = QPushButton("浏览")
        btn_browse_excel.clicked.connect(self._browse_excel)
        file_layout.addWidget(btn_browse_excel)
        excel_layout.addLayout(file_layout)
        
        # 匹配列和数据列
        excel_col_layout = QHBoxLayout()
        excel_col_layout.addWidget(QLabel("匹配列:"))
        self._excel_match_column = QSpinBox()
        self._excel_match_column.setRange(1, 100)
        self._excel_match_column.setValue(element.match_column if element.match_column > 0 else 1)
        self._excel_match_column.setToolTip("图片文件名所在的列号（从1开始），用于匹配对应行")
        excel_col_layout.addWidget(self._excel_match_column)
        
        excel_col_layout.addWidget(QLabel("数据列:"))
        self._excel_data_column = QSpinBox()
        self._excel_data_column.setRange(1, 100)
        self._excel_data_column.setValue(element.data_column if element.data_column > 0 else 2)
        self._excel_data_column.setToolTip("要读取的文本数据列号（从1开始）")
        excel_col_layout.addWidget(self._excel_data_column)
        excel_col_layout.addStretch()
        excel_layout.addLayout(excel_col_layout)
        
        # 起始行
        excel_row_layout = QHBoxLayout()
        excel_row_layout.addWidget(QLabel("数据起始行:"))
        self._excel_row_start = QSpinBox()
        self._excel_row_start.setRange(1, 10000)
        self._excel_row_start.setValue(element.excel_row_start)
        self._excel_row_start.setToolTip("数据从第几行开始（跳过表头）")
        excel_row_layout.addWidget(self._excel_row_start)
        excel_row_layout.addStretch()
        excel_layout.addLayout(excel_row_layout)
        page_layout.addWidget(self._widget_excel)
        
        # Excel使用说明
        excel_hint = QLabel("💡 匹配列：图片文件名所在列 | 数据列：要叠加的文本内容所在列")
        excel_hint.setStyleSheet("color:#666e88;font-size:11px;font-style:italic;")
        excel_hint.setWordWrap(True)
        excel_layout.addWidget(excel_hint)

        # 图片名提示
        self._widget_filename_hint = QLabel("💡 将使用当前处理图片的文件名（不含扩展名）作为文本内容")
        self._widget_filename_hint.setStyleSheet("color:#666e88;font-size:11px;font-style:italic;")
        page_layout.addWidget(self._widget_filename_hint)

        # 根据数据源显示/隐藏对应面板
        def _on_source_change(idx):
            source = combo_source.itemData(idx)
            self._widget_fixed_text.setVisible(source == 'fixed')
            self._widget_excel.setVisible(source == 'excel')
            self._widget_filename_hint.setVisible(source == 'filename')
            element.source = source
        
        combo_source.currentIndexChanged.connect(_on_source_change)
        _on_source_change(combo_source.currentIndex())

        # 字体设置
        font_layout = QHBoxLayout()
        font_layout.addWidget(QLabel("字体:"))
        self._font_family = QComboBox()
        # 常用字体列表（中文字体优先）
        fonts = ['Microsoft YaHei', 'SimHei', 'SimSun', 'KaiTi', 'FangSong', 'Arial', 'Times New Roman']
        self._font_family.addItems(fonts)
        idx = self._font_family.findText(element.font_family)
        if idx >= 0:
            self._font_family.setCurrentIndex(idx)
        else:
            # 默认使用微软雅黑
            idx = self._font_family.findText('Microsoft YaHei')
            if idx >= 0:
                self._font_family.setCurrentIndex(idx)
        
        # 添加加载字体按钮
        btn_load_font = QPushButton("加载字体")
        btn_load_font.clicked.connect(self._load_font)
        font_layout.addWidget(btn_load_font)
        
        self._font_family.setStyleSheet(config.COMBOBOX_STYLE)
        font_layout.addWidget(self._font_family)
        
        font_layout.addWidget(QLabel("大小:"))
        self._font_size = QSpinBox()
        self._font_size.setRange(8, 200)
        self._font_size.setValue(element.font_size)
        font_layout.addWidget(self._font_size)
        
        self._chk_bold = QCheckBox("加粗")
        self._chk_bold.setChecked(element.bold)
        font_layout.addWidget(self._chk_bold)
        
        font_layout.addStretch()
        page_layout.addLayout(font_layout)

        # 颜色选择
        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel("文字颜色:"))
        self._text_color_btn = QPushButton()
        self._text_color_btn.setFixedSize(36, 26)
        self._text_color_btn.setCursor(Qt.PointingHandCursor)
        self._text_color_btn.clicked.connect(self._pick_text_color)
        color_layout.addWidget(self._text_color_btn)
        self._text_color_lbl = QLabel(element.color)
        self._text_color_lbl.setStyleSheet("color:#b0b8c8;font-size:12px;background:transparent;")
        color_layout.addWidget(self._text_color_lbl)
        color_layout.addStretch()
        page_layout.addLayout(color_layout)
        self._refresh_text_color_btn(element.color)

        # 位置设置
        pos_layout = QHBoxLayout()

        # 先创建坐标 SpinBox（供后续宫格引用）
        self._text_x = QSpinBox()
        self._text_x.setRange(0, 99999)
        self._text_x.setValue(element.x)
        self._text_x.setToolTip("相对于图片左上角的X坐标（像素值）")

        self._text_y = QSpinBox()
        self._text_y.setRange(0, 99999)
        self._text_y.setValue(element.y)
        self._text_y.setToolTip("相对于图片左上角的Y坐标（像素值）")

        # 宫格坐标定位器（行首）
        grid = GridPositionWidget(
            x_spin=self._text_x,
            y_spin=self._text_y,
            processor=self,
            element_type='text'
        )
        pos_layout.addWidget(grid)
        pos_layout.addSpacing(8)
        pos_layout.addWidget(QLabel("X坐标(px):"))
        pos_layout.addWidget(self._text_x)
        pos_layout.addWidget(QLabel("Y坐标(px):"))
        pos_layout.addWidget(self._text_y)
        pos_layout.addStretch()
        page_layout.addLayout(pos_layout)

        # 保存引用用于收集参数
        self._current_element = element
        self._combo_source = combo_source
        
        return page

    def _create_image_config_page(self, element: ImageElement) -> QWidget:
        """创建图片元素配置页面"""
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(8)
        
        # 图片文件选择
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("图片文件:"))
        self._image_file_input = QLineEdit()
        self._image_file_input.setText(element.image_path)
        self._image_file_input.setPlaceholderText("选择要叠加的图片...")
        file_layout.addWidget(self._image_file_input)
        btn_browse = QPushButton("浏览")
        btn_browse.clicked.connect(self._browse_image)
        file_layout.addWidget(btn_browse)
        page_layout.addLayout(file_layout)

        # 先创建坐标和尺寸 SpinBox（供后续宫格引用）
        self._image_x = QSpinBox()
        self._image_x.setRange(0, 99999)
        self._image_x.setValue(element.x)
        self._image_x.setToolTip("相对于底图左上角的X坐标（像素值）")

        self._image_y = QSpinBox()
        self._image_y.setRange(0, 99999)
        self._image_y.setValue(element.y)
        self._image_y.setToolTip("相对于底图左上角的Y坐标（像素值）")

        self._image_width = QSpinBox()
        self._image_width.setRange(1, 99999)
        self._image_width.setValue(element.width)
        self._image_width.setToolTip("叠加图片的宽度（像素值）")

        self._image_height = QSpinBox()
        self._image_height.setRange(1, 99999)
        self._image_height.setValue(element.height)
        self._image_height.setToolTip("叠加图片的高度（像素值）")

        # 位置设置（宫格在行首）
        pos_layout = QHBoxLayout()
        grid = GridPositionWidget(
            x_spin=self._image_x,
            y_spin=self._image_y,
            processor=self,
            element_type='image',
            overlay_w_spin=self._image_width,
            overlay_h_spin=self._image_height
        )
        pos_layout.addWidget(grid)
        pos_layout.addSpacing(8)
        pos_layout.addWidget(QLabel("X坐标(px):"))
        pos_layout.addWidget(self._image_x)
        pos_layout.addWidget(QLabel("Y坐标(px):"))
        pos_layout.addWidget(self._image_y)
        pos_layout.addStretch()
        page_layout.addLayout(pos_layout)

        # 大小设置
        size_layout = QHBoxLayout()
        size_layout.addWidget(QLabel("宽度(px):"))
        size_layout.addWidget(self._image_width)
        size_layout.addWidget(QLabel("高度(px):"))
        size_layout.addWidget(self._image_height)
        size_layout.addStretch()
        page_layout.addLayout(size_layout)

        self._current_element = element
        
        return page

    def _browse_excel(self):
        """浏览选择Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._panel, "选择Excel文件", str(Path.home() / "Desktop"), "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self._excel_file_input.setText(file_path)

    def _browse_image(self):
        """浏览选择图片文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._panel, "选择图片文件", str(Path.home() / "Desktop"),
            "Image Files (*.png *.jpg *.jpeg *.webp *.bmp)"
        )
        if file_path:
            self._image_file_input.setText(file_path)

    def _pick_text_color(self):
        """选择文本颜色"""
        current_color = self._text_color_lbl.text()
        c = QColorDialog.getColor(QColor(current_color), self._panel, "选择文字颜色",
                                  QColorDialog.ShowAlphaChannel)
        if c.isValid():
            color = c.name(QColor.HexArgb) if c.alpha() < 255 else c.name()
            self._refresh_text_color_btn(color)
            self._text_color_lbl.setText(color.upper())

    def _refresh_text_color_btn(self, color):
        """刷新颜色按钮样式"""
        self._text_color_btn.setStyleSheet(
            f"QPushButton{{background:{color};border:2px solid #5a5a6a;border-radius:6px;min-width:36px;min-height:24px;}}"
            f"QPushButton:hover{{border-color:#5b8af5;}}"
        )

    def _collect_current_element_options(self):
        """收集当前编辑元素的参数"""
        if not hasattr(self, '_current_element'):
            return
        
        element = self._current_element
        
        if element.element_type == 'text':
            if hasattr(self, '_combo_source'):
                element.source = self._combo_source.currentData()
            if element.source == 'fixed' and hasattr(self, '_text_content'):
                element.content = self._text_content.toPlainText()
            elif element.source == 'excel':
                if hasattr(self, '_excel_file_input'):
                    element.excel_file = self._excel_file_input.text()
                if hasattr(self, '_excel_match_column'):
                    element.match_column = self._excel_match_column.value()
                if hasattr(self, '_excel_data_column'):
                    element.data_column = self._excel_data_column.value()
                if hasattr(self, '_excel_row_start'):
                    element.excel_row_start = self._excel_row_start.value()
            if hasattr(self, '_font_family'):
                element.font_family = self._font_family.currentText()
            if hasattr(self, '_font_size'):
                element.font_size = self._font_size.value()
            if hasattr(self, '_chk_bold'):
                element.bold = self._chk_bold.isChecked()
            if hasattr(self, '_text_color_lbl'):
                element.color = self._text_color_lbl.text()
            if hasattr(self, '_text_x'):
                element.x = self._text_x.value()
            if hasattr(self, '_text_y'):
                element.y = self._text_y.value()
        else:
            if hasattr(self, '_image_file_input'):
                element.image_path = self._image_file_input.text()
            if hasattr(self, '_image_x'):
                element.x = self._image_x.value()
            if hasattr(self, '_image_y'):
                element.y = self._image_y.value()
            if hasattr(self, '_image_width'):
                element.width = self._image_width.value()
            if hasattr(self, '_image_height'):
                element.height = self._image_height.value()

    def gather_options(self) -> dict:
        """收集所有参数"""
        # 先收集当前正在编辑的元素
        self._collect_current_element_options()
        
        # 序列化所有元素
        elements_data = []
        for elem in self._elements:
            if elem.element_type == 'text':
                elements_data.append({
                    'type': 'text',
                    'name': elem.name,
                    'source': elem.source,
                    'content': elem.content,
                    'font_size': elem.font_size,
                    'font_family': elem.font_family,
                    'bold': elem.bold,
                    'color': elem.color,
                    'x': elem.x,
                    'y': elem.y,
                    'excel_file': elem.excel_file,
                    'match_column': elem.match_column,
                    'data_column': elem.data_column,
                    'excel_row_start': elem.excel_row_start,
                })
            else:
                elements_data.append({
                    'type': 'image',
                    'name': elem.name,
                    'image_path': elem.image_path,
                    'x': elem.x,
                    'y': elem.y,
                    'width': elem.width,
                    'height': elem.height,
                })
        
        # 收集自定义字体信息
        custom_fonts = {}
        if hasattr(self, '_custom_fonts'):
            custom_fonts = self._custom_fonts.copy()
        
        return {
            'elements': elements_data,
            'output_format': self.combo_fmt.currentText(),
            'custom_fonts': custom_fonts,
        }

    def get_output_format(self) -> str:
        return self.combo_fmt.currentText()

    def default_options(self) -> dict:
        return {
            'elements': [],
            'output_format': 'png',
            'custom_fonts': {},
        }

    def apply_options(self, options: dict):
        """应用参数到面板"""
        if self._panel is None:
            return
        
        # 先收集当前编辑元素的参数
        self._collect_current_element_options()
        
        # 清除配置页面，显示提示页面
        self._clear_config_pages()
        self._config_stack.setCurrentIndex(0)
        
        # 清除现有元素
        self._elements.clear()
        self._list_widget.clear()
        
        # 加载元素
        elements_data = options.get('elements', [])
        for data in elements_data:
            if data['type'] == 'text':
                element = TextElement(
                    x=data.get('x', 50),
                    y=data.get('y', 50),
                    source=data.get('source', 'fixed'),
                    content=data.get('content', ''),
                    name=data.get('name', ''),
                    font_size=data.get('font_size', 24),
                    font_family=data.get('font_family', 'Microsoft YaHei'),
                    bold=data.get('bold', False),
                    color=data.get('color', '#FFFFFF'),
                    excel_file=data.get('excel_file', ''),
                    match_column=data.get('match_column', 1),
                    data_column=data.get('data_column', 2),
                    excel_row_start=data.get('excel_row_start', 2),
                )
            else:
                element = ImageElement(
                    x=data.get('x', 100),
                    y=data.get('y', 100),
                    image_path=data.get('image_path', ''),
                    name=data.get('name', ''),
                    width=data.get('width', 200),
                    height=data.get('height', 200),
                )
            
            self._elements.append(element)
            item = ElementListItem(element, len(self._elements) - 1)
            self._list_widget.addItem(item)
        
        # 设置输出格式
        fmt = options.get('output_format', 'png')
        idx = ['png', 'webp', 'jpg'].index(fmt) if fmt in ['png', 'webp', 'jpg'] else 0
        self.combo_fmt.setCurrentIndex(idx)
        
        # 恢复自定义字体
        custom_fonts = options.get('custom_fonts', {})
        if custom_fonts:
            self._custom_fonts = custom_fonts
            # 将自定义字体添加到字体下拉框中
            existing_fonts = [self._font_family.itemText(i) for i in range(self._font_family.count())]
            for font_name in custom_fonts.keys():
                if font_name not in existing_fonts:
                    self._font_family.addItem(font_name)

    def _load_font(self):
        """加载外部字体文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self._panel, "选择字体文件", str(Path.home() / "Desktop"),
            "Font Files (*.ttf *.otf *.ttc)"
        )
        if not file_path:
            return
        
        try:
            # 验证字体文件是否有效
            test_font = ImageFont.truetype(file_path, 12)
            
            # 获取字体名称
            font_name = Path(file_path).stem
            
            # 检查是否已存在
            existing_fonts = [self._font_family.itemText(i) for i in range(self._font_family.count())]
            if font_name in existing_fonts:
                QMessageBox.warning(self._panel, "提示", f"字体 '{font_name}' 已存在")
                return
            
            # 添加到字体列表
            self._font_family.addItem(font_name)
            self._font_family.setCurrentIndex(self._font_family.count() - 1)
            
            # 复制字体文件到程序字体目录（可选，这里我们直接保存路径）
            # 为了简化，我们直接在字体映射中记录路径
            if not hasattr(self, '_custom_fonts'):
                self._custom_fonts = {}
            self._custom_fonts[font_name] = file_path
            
            QMessageBox.information(self._panel, "成功", f"字体 '{font_name}' 加载成功！")
            
        except Exception as e:
            QMessageBox.warning(self._panel, "错误", f"字体加载失败: {e}")

    def _get_font(self, font_family: str, font_size: int, bold: bool):
        """获取字体，支持中文字体，使用缓存避免重复加载"""
        # 生成缓存键
        cache_key = f"{font_family}_{font_size}_{bold}"
        
        # 检查缓存
        if cache_key in self._font_cache:
            return self._font_cache[cache_key]
        
        # 缓存过大时清理（防止内存泄漏）
        if len(self._font_cache) > 100:
            self._clear_font_cache()
        
        try:
            # 尝试使用指定字体
            font_path = None
            
            # 检查是否为自定义字体
            if hasattr(self, '_custom_fonts') and font_family in self._custom_fonts:
                font_path = self._custom_fonts[font_family]
                font = ImageFont.truetype(font_path, font_size)
                self._font_cache[cache_key] = font
                return font
            
            # Windows 系统字体路径
            if os.name == 'nt':
                font_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
                
                # 字体映射（优先使用中文字体）
                font_map = {
                    'Microsoft YaHei': ('msyh.ttc', 0),  # 微软雅黑，索引0为常规体
                    'SimHei': ('simhei.ttf', 0),         # 黑体
                    'SimSun': ('simsun.ttc', 0),         # 宋体，索引0
                    'KaiTi': ('simkai.ttf', 0),          # 楷体
                    'FangSong': ('simfang.ttf', 0),      # 仿宋
                    'Arial': ('arial.ttf', 0),           # Arial（不支持中文）
                    'Times New Roman': ('times.ttf', 0), # Times New Roman（不支持中文）
                }
                
                font_info = font_map.get(font_family)
                if font_info:
                    font_file, font_index = font_info
                    font_path = os.path.join(font_dir, font_file)
                    if not os.path.exists(font_path):
                        font_path = None
                    
                    # 如果是 .ttc 字体文件，需要指定索引
                    if font_path and font_path.endswith('.ttc'):
                        font = ImageFont.truetype(font_path, font_size, index=font_index)
                        self._font_cache[cache_key] = font
                        return font
                    elif font_path:
                        font = ImageFont.truetype(font_path, font_size)
                        self._font_cache[cache_key] = font
                        return font
            
            # 如果指定字体加载失败，尝试加载中文字体作为回退
            fallback_fonts = ['msyh.ttc', 'simhei.ttf', 'simsun.ttc']
            if os.name == 'nt':
                font_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
                for fallback in fallback_fonts:
                    fallback_path = os.path.join(font_dir, fallback)
                    if os.path.exists(fallback_path):
                        if fallback.endswith('.ttc'):
                            font = ImageFont.truetype(fallback_path, font_size, index=0)
                            self._font_cache[cache_key] = font
                            return font
                        else:
                            font = ImageFont.truetype(fallback_path, font_size)
                            self._font_cache[cache_key] = font
                            return font
            
            # 最终回退到默认字体
            font = ImageFont.load_default()
            self._font_cache[cache_key] = font
            return font
        except Exception as e:
            # 异常时回退到默认字体
            print(f"字体加载失败: {e}，使用默认字体")
            font = ImageFont.load_default()
            self._font_cache[cache_key] = font
            return font

    def _read_excel_data(self, excel_file: str, match_column: int, data_column: int, 
                         row_start: int, image_filename: str):
        """从Excel读取数据，通过文件名匹配对应行"""
        try:
            import openpyxl
            
            if not excel_file or not os.path.exists(excel_file):
                return ""
            
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # 获取文件名（不含扩展名）
            image_stem = Path(image_filename).stem
            
            # 遍历数据行，查找匹配的文件名
            for row in range(row_start, ws.max_row + 1):
                # 读取匹配列的值
                match_col_letter = openpyxl.utils.get_column_letter(match_column)
                match_value = ws[f'{match_col_letter}{row}'].value
                
                if match_value is None:
                    continue
                
                # 转换为字符串并去除扩展名（如果匹配列包含扩展名）
                match_str = str(match_value).strip()
                # 去除可能的扩展名
                if '.' in match_str:
                    match_stem = Path(match_str).stem
                else:
                    match_stem = match_str
                
                # 匹配文件名（不区分大小写）
                if match_stem.lower() == image_stem.lower():
                    # 找到匹配行，读取数据列
                    data_col_letter = openpyxl.utils.get_column_letter(data_column)
                    data_value = ws[f'{data_col_letter}{row}'].value
                    return str(data_value) if data_value is not None else ""
            
            # 未找到匹配行
            return ""
            
        except Exception as e:
            return f"[Excel读取错误: {str(e)}]"

    def _get_text_content(self, element: TextElement, image_path: str, image_index: int):
        """获取文本内容"""
        if element.source == 'fixed':
            return element.content
        elif element.source == 'filename':
            # 获取文件名（不含扩展名）
            return Path(image_path).stem
        elif element.source == 'excel':
            return self._read_excel_data(
                element.excel_file, 
                element.match_column,
                element.data_column,
                element.excel_row_start,
                image_path
            )
        return ""

    def process(self, img: Image.Image, options: dict) -> tuple[Image.Image, dict]:
        """处理单张图片，叠加元素"""
        details = {"original_size": img.size, "overlays": []}
        
        # 转换为RGBA
        img = img.convert("RGBA")
        img_width, img_height = img.size
        
        # 注意：process方法在批量处理时会被多次调用
        # 我们需要知道当前是第几张图片，以便正确读取Excel数据
        # 但BaseProcessor的process接口没有提供image_index参数
        # 这里我们使用一个临时方案：从options中获取（需要在worker中传递）
        image_index = options.get('_image_index', 0)
        current_image_path = options.get('_current_image_path', '')
        
        # 创建绘图对象
        draw = ImageDraw.Draw(img)
        
        elements_data = options.get('elements', [])
        
        for idx, elem_data in enumerate(elements_data):
            try:
                if elem_data['type'] == 'text':
                    # 文本元素
                    text_content = ""
                    if elem_data['source'] == 'fixed':
                        text_content = elem_data.get('content', '')
                    elif elem_data['source'] == 'filename':
                        text_content = Path(current_image_path).stem if current_image_path else ""
                    elif elem_data['source'] == 'excel':
                        text_content = self._read_excel_data(
                            elem_data.get('excel_file', ''),
                            elem_data.get('match_column', 1),
                            elem_data.get('data_column', 2),
                            elem_data.get('excel_row_start', 2),
                            current_image_path
                        )
                    
                    if not text_content:
                        details["overlays"].append({
                            "type": "text_skipped",
                            "index": idx,
                            "reason": "文本内容为空",
                            "source": elem_data.get('source', 'unknown'),
                            "position": (elem_data['x'], elem_data['y'])
                        })
                        continue
                    
                    # 直接使用像素坐标
                    x = elem_data['x']
                    y = elem_data['y']
                    
                    # 获取字体
                    font_family = elem_data.get('font_family', 'Microsoft YaHei')
                    font_size = elem_data.get('font_size', 24)
                    bold = elem_data.get('bold', False)
                    font = self._get_font(font_family, font_size, bold)
                    
                    # 获取颜色
                    color = hex_to_rgba(elem_data.get('color', '#FFFFFF'))
                    
                    # 绘制文本
                    draw.text((x, y), text_content, fill=color, font=font)
                    
                    details["overlays"].append({
                        "type": "text",
                        "index": idx,
                        "content": text_content[:30],
                        "font": f"{font_family} {font_size}px",
                        "position": (x, y),
                        "color": elem_data.get('color', '#FFFFFF')
                    })
                
                elif elem_data['type'] == 'image':
                    # 图片元素
                    overlay_path = elem_data.get('image_path', '')
                    if not overlay_path or not os.path.exists(overlay_path):
                        continue
                    
                    # 直接使用像素坐标和尺寸
                    x = elem_data['x']
                    y = elem_data['y']
                    overlay_w = elem_data['width']
                    overlay_h = elem_data['height']
                    
                    # 加载叠加图片
                    overlay_img = Image.open(overlay_path).convert("RGBA")
                    overlay_img = overlay_img.resize((overlay_w, overlay_h), Image.LANCZOS)
                    
                    # 粘贴到目标位置
                    img.paste(overlay_img, (x, y), overlay_img)
                    
                    # 关闭叠加图片释放内存
                    overlay_img.close()
                    
                    details["overlays"].append({
                        "type": "image",
                        "index": idx,
                        "path": Path(overlay_path).name,
                        "position": (x, y),
                        "size": (overlay_w, overlay_h)
                    })
            
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()
                details["overlays"].append({
                    "type": "error",
                    "index": idx,
                    "error": str(e),
                    "traceback": error_trace[:200]  # 只保留前200字符
                })
        
        # 记录处理统计
        text_count = sum(1 for o in details["overlays"] if o.get("type") == "text")
        image_count = sum(1 for o in details["overlays"] if o.get("type") == "image")
        skipped_count = sum(1 for o in details["overlays"] if o.get("type") in ("text_skipped", "image_skipped"))
        error_count = sum(1 for o in details["overlays"] if o.get("type") == "error")
        details["summary"] = f"文本:{text_count}, 图片:{image_count}, 跳过:{skipped_count}, 错误:{error_count}, 总计:{len(elements_data)}"
        
        # 清理绘图对象
        del draw
        
        return img, details
